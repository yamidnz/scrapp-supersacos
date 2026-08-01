"""
Sincroniza el archivo Excel de control de scrap desde SharePoint hacia data/scrap-data.json.

Se ejecuta dentro de un workflow de GitHub Actions. Usa el flujo "client credentials"
de Microsoft Graph, así que NO requiere que nadie inicie sesión ni deje una sesión abierta.
Todas las credenciales llegan como variables de entorno (configuradas como Secrets en GitHub,
nunca quedan escritas en este archivo ni en el repositorio).

Variables de entorno requeridas:
  MS_TENANT_ID       - ID del tenant de Microsoft 365 (Directory (tenant) ID)
  MS_CLIENT_ID       - ID de la app registration (Application (client) ID)
  MS_CLIENT_SECRET   - El secreto generado para esa app registration
  SP_SITE_HOSTNAME   - Ej: tuempresa.sharepoint.com
  SP_SITE_PATH       - Ej: /sites/SuperSacos  (déjalo vacío "" si el archivo está en el sitio raíz)
  SP_FILE_PATH       - Ruta del archivo dentro de la biblioteca de documentos,
                        Ej: /Documentos compartidos/SCRAPP SS.xlsx
  SP_SHEET_NAME       - (Opcional) nombre exacto de la hoja a leer. Si se deja vacío,
                        el script elige automáticamente la hoja con más filas de datos.
"""

import os
import re
import sys
import json
import unicodedata
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook

TENANT_ID = os.environ["MS_TENANT_ID"]
CLIENT_ID = os.environ["MS_CLIENT_ID"]
CLIENT_SECRET = os.environ["MS_CLIENT_SECRET"]
SITE_HOSTNAME = os.environ["SP_SITE_HOSTNAME"]
SITE_PATH = os.environ.get("SP_SITE_PATH", "").strip()
FILE_PATH = os.environ["SP_FILE_PATH"]
SHEET_NAME = os.environ.get("SP_SHEET_NAME", "").strip()

GRAPH = "https://graph.microsoft.com/v1.0"


def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }
    r = requests.post(url, data=data, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


def get_site_id(token):
    site_ref = f"{SITE_HOSTNAME}:{SITE_PATH}" if SITE_PATH else SITE_HOSTNAME
    url = f"{GRAPH}/sites/{site_ref}"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    r.raise_for_status()
    return r.json()["id"]


def download_file(token, site_id):
    path = FILE_PATH if FILE_PATH.startswith("/") else "/" + FILE_PATH
    url = f"{GRAPH}/sites/{site_id}/drive/root:{path}:/content"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    return r.content


def normalize(s):
    s = str(s or "").lower().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def guess_col(headers, keywords):
    for h in headers:
        n = normalize(h)
        if any(k in n for k in keywords):
            return h
    return None


def pick_sheet(wb):
    if SHEET_NAME and SHEET_NAME in wb.sheetnames:
        return SHEET_NAME
    best_name, best_rows = wb.sheetnames[0], -1
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row or 0
        if rows > best_rows:
            best_rows, best_name = rows, name
    return best_name


def main():
    print("Autenticando contra Microsoft Graph...")
    token = get_token()

    print("Resolviendo el sitio de SharePoint...")
    site_id = get_site_id(token)

    print(f"Descargando archivo: {FILE_PATH}")
    file_bytes = download_file(token, site_id)

    with open("_tmp_source.xlsx", "wb") as f:
        f.write(file_bytes)

    wb = load_workbook("_tmp_source.xlsx", data_only=True)
    sheet_name = pick_sheet(wb)
    ws = wb[sheet_name]
    print(f"Leyendo hoja: {sheet_name}")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    col_semana = guess_col(headers, ["seman", "week"])
    col_categoria = guess_col(headers, ["categ"])
    col_subcategoria = guess_col(headers, ["subcat"])
    col_peso = guess_col(headers, ["peso", "kg"])
    col_fecha = guess_col(headers, ["fecha", "date"])

    # Evita que categoría y subcategoría se detecten como la misma columna
    if col_categoria and col_categoria == col_subcategoria:
        candidates = [h for h in headers if "categ" in normalize(h) and h != col_subcategoria]
        col_categoria = candidates[0] if candidates else col_categoria

    missing = [n for n, v in [("semana", col_semana), ("categoria", col_categoria),
                               ("subcategoria", col_subcategoria), ("peso", col_peso)] if not v]
    if missing:
        print(f"ERROR: no se detectaron las columnas: {missing}. Encabezados encontrados: {headers}")
        sys.exit(1)

    idx = {h: i for i, h in enumerate(headers)}
    out_rows = []
    for row in rows_iter:
        if row is None:
            continue

        def cell(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        semana = str(cell(col_semana) or "").strip()
        categoria = str(cell(col_categoria) or "").strip().upper()
        subcategoria = str(cell(col_subcategoria) or "").strip().upper()
        peso_raw = cell(col_peso)
        try:
            peso = float(peso_raw) if peso_raw is not None else 0.0
        except (TypeError, ValueError):
            peso = 0.0
        fecha_raw = cell(col_fecha) if col_fecha else None
        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw.strftime("%Y-%m-%d")
        elif fecha_raw:
            fecha = str(fecha_raw).strip()
        else:
            fecha = ""

        if not semana or not categoria:
            continue
        out_rows.append({
            "semana": semana,
            "categoria": categoria,
            "subcategoria": subcategoria,
            "fecha": fecha,
            "peso": peso,
        })

    print(f"Filas convertidas: {len(out_rows)}")

    os.makedirs("data", exist_ok=True)
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_sheet": sheet_name,
        "rows": out_rows,
    }
    with open("data/scrap-data.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)

    os.remove("_tmp_source.xlsx")
    print("Listo: data/scrap-data.json actualizado.")


if __name__ == "__main__":
    main()
